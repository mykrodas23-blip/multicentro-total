import streamlit as st
import pandas as pd
from datetime import datetime
import io
import os
import urllib.parse
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as OpenPyxlImage

# Configuración de página
st.set_page_config(
    page_title="Multicentro Total - Punto de Venta",
    page_icon="🏪",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Estilos CSS
st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .stApp { background-color: #f0f0f0; }
    .top-title { font-weight: bold; font-size: 20px; color: #1a365d; }
    .sales-banner {
        background-color: #3b82f6; color: white;
        padding: 6px 12px; font-weight: bold; font-size: 16px;
        border-radius: 3px 3px 0 0; margin-top: 5px;
    }
    .total-display {
        background-color: #ffffff; border: 2px solid #cbd5e1;
        border-radius: 8px; padding: 10px; text-align: right;
    }
    .total-amount { font-size: 42px; font-weight: bold; color: #0000d1; }
    .stButton>button { width: 100%; border-radius: 4px; font-weight: 600; }
    </style>
""", unsafe_allow_html=True)

# Listado de tiendas
TIENDAS = [
    "Multicentro Total (Aldea Chuatzunuj)",
    "Multicentro Total 2 (Aldea Caquixajay)"
]

# LISTA OFICIAL DE ÚNICAS ÁREAS DE PRODUCTO
AREAS_OFICIALES = [
    "Ferretería",
    "Abarrotes",
    "Librería",
    "Farmacia",
    "Heladería"
]

CELULAR_FRANCISCO = "50236372449"

# --- INICIALIZACIÓN DE ESTADO ---
if 'usuarios' not in st.session_state:
    st.session_state.usuarios = {
        "Francisco": {"clave": "123456", "rol": "Administrador", "tienda": "Todas", "creado_por": "Sistema"},
        "Francisco Sanic": {"clave": "777*123", "rol": "Administrador", "tienda": "Todas", "creado_por": "Sistema"},
        "Vendedor1": {"clave": "0000", "rol": "Vendedor de tienda", "tienda": "Multicentro Total (Aldea Chuatzunuj)", "creado_por": "Francisco Sanic"},
        "Vendedor2": {"clave": "0000", "rol": "Vendedor de tienda", "tienda": "Multicentro Total 2 (Aldea Caquixajay)", "creado_por": "Francisco Sanic"}
    }

if 'usuario_logueado' not in st.session_state:
    st.session_state.usuario_logueado = None

if 'rol_logueado' not in st.session_state:
    st.session_state.rol_logueado = None

if 'tienda_asignada_user' not in st.session_state:
    st.session_state.tienda_asignada_user = None

if 'tienda_activa' not in st.session_state:
    st.session_state.tienda_activa = TIENDAS[0]

if 'carrito' not in st.session_state:
    st.session_state.carrito = []

if 'tickets_pendientes' not in st.session_state:
    st.session_state.tickets_pendientes = []

if 'ultimo_ticket' not in st.session_state:
    st.session_state.ultimo_ticket = None

if 'vista_actual' not in st.session_state:
    st.session_state.vista_actual = "ventas"

if 'alerta_stock_oculta' not in st.session_state:
    st.session_state.alerta_stock_oculta = False

if 'ver_detalle_stock' not in st.session_state:
    st.session_state.ver_detalle_stock = False

if 'ventas_realizadas' not in st.session_state:
    st.session_state.ventas_realizadas = [
        {"fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "tienda": TIENDAS[0], "vendedor": "Vendedor1", "pago": "Efectivo", "total": 18.50, "costo_total": 14.00, "items": 1},
        {"fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "tienda": TIENDAS[0], "vendedor": "Francisco Sanic", "pago": "Tarjeta", "total": 45.00, "costo_total": 32.00, "items": 10}
    ]

if 'proveedores' not in st.session_state:
    st.session_state.proveedores = [
        {
            "id": 1, 
            "empresa": "Distribuidora El Sol", 
            "propietario": "Carlos López", 
            "telefono": "502 5555 1234", 
            "direccion": "Cantón Central, Sololá", 
            "area": "Abarrotes", 
            "notas": "Entregas los martes"
        },
        {
            "id": 2, 
            "empresa": "Ferretería Central S.A.", 
            "propietario": "Ing. Mario Gómez", 
            "telefono": "502 4444 8888", 
            "direccion": "Zona 4, Ciudad de Guatemala", 
            "area": "Ferretería", 
            "notas": "Crédito a 15 días"
        }
    ]

if 'movimientos_inventario' not in st.session_state:
    st.session_state.movimientos_inventario = []

if 'mensaje_cobro' not in st.session_state:
    st.session_state.mensaje_cobro = None

if 'mensaje_bienvenida' not in st.session_state:
    st.session_state.mensaje_bienvenida = None

if 'mensaje_prod_accion' not in st.session_state:
    st.session_state.mensaje_prod_accion = None

if 'mensaje_user_accion' not in st.session_state:
    st.session_state.mensaje_user_accion = None

if 'mensaje_prov_accion' not in st.session_state:
    st.session_state.mensaje_prov_accion = None

if 'confirmar_borrado_cod' not in st.session_state:
    st.session_state.confirmar_borrado_cod = None

if 'confirmar_borrado_user' not in st.session_state:
    st.session_state.confirmar_borrado_user = None

# Inventarios por tienda
if 'inventario_tiendas' not in st.session_state:
    st.session_state.inventario_tiendas = {
        "Multicentro Total (Aldea Chuatzunuj)": {
            "75010001": {"nombre": "Aceite Vegetal 1L", "area": "Abarrotes", "precio": 18.50, "costo": 14.00, "existencia": 3.0, "stock_minimo": 5.0, "proveedor": "Distribuidora El Sol", "foto": None, "creado_por": "Sistema Base", "fecha_registro": "2026-07-01 08:00"},
            "75010002": {"nombre": "Cuaderno Espiral 100 Hojas", "area": "Librería", "precio": 12.00, "costo": 8.50, "existencia": 50.0, "stock_minimo": 10.0, "proveedor": "Distribuidora El Sol", "foto": None, "creado_por": "Sistema Base", "fecha_registro": "2026-07-01 08:00"},
            "FER-01": {"nombre": "Martillo de Uña 16oz", "area": "Ferretería", "precio": 45.00, "costo": 32.00, "existencia": 2.0, "stock_minimo": 3.0, "proveedor": "Ferretería Central S.A.", "foto": None, "creado_por": "Sistema Base", "fecha_registro": "2026-07-01 08:00"},
            "FAR-01": {"nombre": "Paracetamol 500mg (Blíster)", "area": "Farmacia", "precio": 10.00, "costo": 6.00, "existencia": 80.0, "stock_minimo": 15.0, "proveedor": "Sin Proveedor", "foto": None, "creado_por": "Sistema Base", "fecha_registro": "2026-07-01 08:00"},
            "HEL-01": {"nombre": "Helado de Sombrilla Varios Sabores", "area": "Heladería", "precio": 5.00, "costo": 2.50, "existencia": 40.0, "stock_minimo": 10.0, "proveedor": "Sin Proveedor", "foto": None, "creado_por": "Sistema Base", "fecha_registro": "2026-07-01 08:00"}
        },
        "Multicentro Total 2 (Aldea Caquixajay)": {
            "75010001": {"nombre": "Aceite Vegetal 1L", "area": "Abarrotes", "precio": 18.50, "costo": 14.00, "existencia": 10.0, "stock_minimo": 5.0, "proveedor": "Distribuidora El Sol", "foto": None, "creado_por": "Sistema Base", "fecha_registro": "2026-07-01 08:00"},
            "HEL-01": {"nombre": "Helado de Sombrilla Varios Sabores", "area": "Heladería", "precio": 5.00, "costo": 2.50, "existencia": 4.0, "stock_minimo": 5.0, "proveedor": "Sin Proveedor", "foto": None, "creado_por": "Sistema Base", "fecha_registro": "2026-07-01 08:00"}
        }
    }

# --- PANTALLA DE LOGIN ---
if st.session_state.usuario_logueado is None:
    st.write("<br>", unsafe_allow_html=True)
    col_l1, col_l2, col_l3 = st.columns([1, 1.8, 1])
    
    with col_l2:
        if os.path.exists('logo.png'):
            col_img1, col_img2, col_img3 = st.columns([1, 2, 1])
            with col_img2:
                st.image('logo.png', width=220)
        else:
            st.markdown("<h2 style='text-align: center; color: #1a365d;'>🏪 Multicentro Total</h2>", unsafe_allow_html=True)
        
        st.markdown("<h4 style='text-align: center;'>Punto de Venta</h4>", unsafe_allow_html=True)
        
        with st.form("form_login"):
            usuario_input = st.text_input("Usuario")
            clave_input = st.text_input("Contraseña", type="password")
            btn_ingresar = st.form_submit_button("Ingresar al Sistema", type="primary")
            
            if btn_ingresar:
                if usuario_input in st.session_state.usuarios and st.session_state.usuarios[usuario_input]["clave"] == clave_input:
                    u_data = st.session_state.usuarios[usuario_input]
                    st.session_state.usuario_logueado = usuario_input
                    st.session_state.rol_logueado = u_data["rol"]
                    st.session_state.tienda_asignada_user = u_data.get("tienda", "Todas")
                    
                    if u_data["rol"] == "Vendedor de tienda" and u_data.get("tienda") in TIENDAS:
                        st.session_state.tienda_activa = u_data["tienda"]
                    else:
                        st.session_state.tienda_activa = TIENDAS[0]
                    
                    st.session_state.alerta_stock_oculta = False
                    st.session_state.ver_detalle_stock = False
                    st.session_state.mensaje_bienvenida = f"👋 ¡Bienvenido al sistema, **{usuario_input}**! Has ingresado a **{st.session_state.tienda_activa}**."
                    st.session_state.vista_actual = "ventas"
                    st.rerun()
                else:
                    st.error("❌ Usuario o contraseña incorrectos.")
    st.stop()

# --- HEADER SUPERIOR (Con Botón de Logo / Inicio) ---
col_logo, col_atendido = st.columns([2.5, 1.5])

with col_logo:
    if st.button("🏠 Ir a Inicio / Ventas", key="btn_logo_home", use_container_width=False):
        st.session_state.vista_actual = "ventas"
        st.session_state.mensaje_cobro = None
        st.rerun()
        
    try:
        if os.path.exists('logo.png'):
            st.image('logo.png', width=190)
        else:
            st.markdown("<div class='top-title'>🏪 <b>Multicentro Total</b> - Punto de Venta</div>", unsafe_allow_html=True)
    except:
        st.markdown("<div class='top-title'>🏪 <b>Multicentro Total</b> - Punto de Venta</div>", unsafe_allow_html=True)

with col_atendido:
    st.caption(f"👤 **Usuario:** {st.session_state.usuario_logueado} ({st.session_state.rol_logueado})")
    
    if st.session_state.rol_logueado == "Administrador" or st.session_state.tienda_asignada_user == "Ambas Tiendas":
        tienda_sel = st.selectbox(
            "📍 **Tienda Gestionada:**",
            TIENDAS,
            index=TIENDAS.index(st.session_state.tienda_activa) if st.session_state.tienda_activa in TIENDAS else 0,
            key="sel_tienda_header"
        )
        if tienda_sel != st.session_state.tienda_activa:
            st.session_state.tienda_activa = tienda_sel
            st.session_state.carrito = []
            st.session_state.alerta_stock_oculta = False
            st.session_state.ver_detalle_stock = False
            st.session_state.mensaje_cobro = None
            st.session_state.mensaje_bienvenida = None
            st.session_state.mensaje_prod_accion = None
            st.rerun()
    else:
        st.info(f"📍 **Tienda:** {st.session_state.tienda_activa}")

    if st.button("🚪 Cerrar Sesión", key="btn_logout"):
        st.session_state.usuario_logueado = None
        st.session_state.rol_logueado = None
        st.session_state.tienda_asignada_user = None
        st.session_state.carrito = []
        st.session_state.alerta_stock_oculta = False
        st.session_state.ver_detalle_stock = False
        st.session_state.mensaje_cobro = None
        st.session_state.mensaje_bienvenida = None
        st.session_state.mensaje_prod_accion = None
        st.session_state.mensaje_user_accion = None
        st.session_state.vista_actual = "ventas"
        st.rerun()

inventario_actual = st.session_state.inventario_tiendas[st.session_state.tienda_activa]

# Verificación de alertas de Stock Mínimo
prods_agotandose = [
    {"codigo": k, "nombre": v['nombre'], "area": v['area'], "existencia": v['existencia'], "stock_minimo": v['stock_minimo']}
    for k, v in inventario_actual.items()
    if v['existencia'] <= v['stock_minimo']
]

# --- BARRA DE NAVEGACIÓN ---
nav_cols = st.columns(8)

btn_ventas = nav_cols[0].button("🛒 Ventas", type="primary" if st.session_state.vista_actual == "ventas" else "secondary", key="nav_v")
btn_clientes = nav_cols[1].button("👥 Catálogos", type="primary" if st.session_state.vista_actual == "clientes" else "secondary", key="nav_c")
btn_productos = nav_cols[2].button("🏷️ Productos", type="primary" if st.session_state.vista_actual == "productos" else "secondary", key="nav_p")
btn_inventario = nav_cols[3].button("📦 Inventario", type="primary" if st.session_state.vista_actual == "inventario" else "secondary", key="nav_i")

btn_proveedores = nav_cols[4].button("🚛 Proveedores", type="primary" if st.session_state.vista_actual == "proveedores" else "secondary", key="nav_prov")

es_admin = (st.session_state.rol_logueado == "Administrador")

btn_ganancias = nav_cols[5].button(
    "📈 Ganancias",
    type="primary" if st.session_state.vista_actual == "ganancias" else "secondary",
    key="nav_gan",
    disabled=not es_admin,
    help="Reportes de utilidades (Administradores)" if es_admin else "Acceso restringido"
)

btn_config = nav_cols[6].button(
    "⚙️ Config", 
    type="primary" if st.session_state.vista_actual == "config" else "secondary", 
    key="nav_cfg",
    disabled=not es_admin,
    help="Gestión de usuarios" if es_admin else "Acceso restringido"
)

btn_corte = nav_cols[7].button("📊 Corte", type="primary" if st.session_state.vista_actual == "corte" else "secondary", key="nav_crt")

if btn_ventas: st.session_state.vista_actual = "ventas"; st.rerun()
if btn_clientes: st.session_state.vista_actual = "clientes"; st.rerun()
if btn_productos: st.session_state.vista_actual = "productos"; st.rerun()
if btn_inventario: st.session_state.vista_actual = "inventario"; st.rerun()
if btn_proveedores: st.session_state.vista_actual = "proveedores"; st.rerun()
if btn_ganancias and es_admin: st.session_state.vista_actual = "ganancias"; st.rerun()
if btn_config and es_admin: st.session_state.vista_actual = "config"; st.rerun()
if btn_corte: st.session_state.vista_actual = "corte"; st.rerun()

# --- BANNER INTELIGENTE DE ALERTA DE STOCK MÍNIMO ---
if prods_agotandose and not st.session_state.alerta_stock_oculta:
    textos_wa = [f"• {p['nombre']} ({p['area']}): Stock actual {p['existencia']:.1f} (Mínimo: {p['stock_minimo']:.1f})" for p in prods_agotandose]
    msg_wa = f"Hola Francisco, te saluda {st.session_state.usuario_logueado} de {st.session_state.tienda_activa}.\n\n*ALERTA DE STOCK MÍNIMO REQUERIDO:*\n" + "\n".join(textos_wa)
    url_wa = f"https://wa.me/{CELULAR_FRANCISCO}?text={urllib.parse.quote(msg_wa)}"

    if es_admin:
        col_alt1, col_alt2, col_alt3, col_alt4 = st.columns([2.4, 1.1, 1.1, 0.7])
    else:
        col_alt1, col_alt2, col_alt3 = st.columns([3, 1.2, 0.8])
    
    with col_alt1:
        st.warning(f"⚠️ **ALERTA DE STOCK BAJO EN {st.session_state.tienda_activa.upper()}:** Hay {len(prods_agotandose)} producto(s) por debajo del mínimo.")
    
    if es_admin:
        with col_alt2:
            txt_btn_det = "🔼 Ocultar Detalle" if st.session_state.ver_detalle_stock else "🔍 Ver Artículos Bajos"
            if st.button(txt_btn_det, key="btn_toggle_detalle_stock"):
                st.session_state.ver_detalle_stock = not st.session_state.ver_detalle_stock
                st.rerun()

    with col_alt2 if not es_admin else col_alt3:
        st.markdown(
            f'''
            <a href="{url_wa}" target="_blank" style="
                display: block;
                background-color: #25D366;
                color: white;
                padding: 0.45rem 1rem;
                text-align: center;
                text-decoration: none;
                font-weight: 600;
                font-size: 14px;
                border-radius: 4px;
                border: none;
            ">
                📲 WhatsApp a Francisco
            </a>
            ''',
            unsafe_allow_html=True
        )

    with col_alt3 if not es_admin else col_alt4:
        if st.button("✖️ Ocultar", key="btn_dismiss_alert"):
            st.session_state.alerta_stock_oculta = True
            st.session_state.ver_detalle_stock = False
            st.rerun()

    if es_admin and st.session_state.ver_detalle_stock:
        st.info("📋 **Detalle de Artículos por debajo del Stock Mínimo:**")
        df_stock_bajo = pd.DataFrame(prods_agotandose)[["codigo", "nombre", "area", "existencia", "stock_minimo"]]
        df_stock_bajo.columns = ["Código", "Descripción", "Área", "Existencia Actual", "Stock Mínimo"]
        st.dataframe(df_stock_bajo, use_container_width=True, hide_index=True)

# --- VISTA: VENTAS ---
if st.session_state.vista_actual == "ventas":
    st.markdown(f"<div class='sales-banner'>VENTA EN LÍNEA - {st.session_state.tienda_activa}</div>", unsafe_allow_html=True)

    if st.session_state.mensaje_bienvenida:
        st.info(st.session_state.mensaje_bienvenida)
        st.session_state.mensaje_bienvenida = None

    if st.session_state.mensaje_cobro:
        st.success(st.session_state.mensaje_cobro)

    def agregar_producto_por_codigo(codigo, cantidad=1.0):
        if codigo in inventario_actual:
            item = inventario_actual[codigo]
            encontrado = False
            for prod in st.session_state.carrito:
                if prod['codigo'] == codigo:
                    prod['cant'] += cantidad
                    prod['importe'] = prod['cant'] * prod['precio']
                    prod['costo_subtotal'] = prod['cant'] * item.get('costo', 0.0)
                    encontrado = True
                    break
            if not encontrado:
                st.session_state.carrito.append({
                    "codigo": codigo,
                    "descripcion": item['nombre'],
                    "precio": item['precio'],
                    "costo": item.get('costo', 0.0),
                    "cant": cantidad,
                    "importe": cantidad * item['precio'],
                    "costo_subtotal": cantidad * item.get('costo', 0.0),
                    "existencia": item['existencia']
                })
            st.session_state.mensaje_cobro = None

    col_busq, col_cant, col_pago = st.columns([2.5, 1, 1.2])
    
    with col_cant:
        cant_venta = st.number_input("Cantidad a Ingresar:", min_value=0.01, value=1.0, step=1.0, format="%.2f", key="input_cant_venta", help="Indica cuántas unidades deseas agregar antes de buscar o escanear el código.")

    with col_pago:
        metodo_pago = st.selectbox("Método de Pago:", ["Efectivo", "Transferencia", "Tarjeta"], key="metodo_pago_sel")

    with col_busq:
        busqueda = st.text_input("🔎 Escanear Código o Escribir (Ej. 'Cla'):", placeholder="Escribe las primeras letras...", key="busqueda_input").strip()

    # BÚSQUEDA PREDICTIVA INTERACTIVA
    if busqueda:
        coincidencias = {
            k: v for k, v in inventario_actual.items() 
            if busqueda.lower() in k.lower() or busqueda.lower() in v['nombre'].lower()
        }
        
        if len(coincidencias) == 1:
            cod_nico = list(coincidencias.keys())[0]
            prod_unico = coincidencias[cod_nico]
            if st.button(f"➕ Agregar {cant_venta} x {prod_unico['nombre']} (Q{prod_unico['precio']:.2f})", key="btn_add_predictivo_unico", type="primary"):
                agregar_producto_por_codigo(cod_nico, cant_venta)
                st.rerun()
        elif len(coincidencias) > 1:
            st.markdown("💡 **Coincidencias encontradas (Haz clic en el producto para agregarlo):**")
            opciones_sugeridas = {f"{v['nombre']} | Precio: Q{v['precio']:.2f} | Stock: {v['existencia']:.1f} ({k})": k for k, v in coincidencias.items()}
            prod_elegido_label = st.selectbox("Selecciona de la lista desplegable:", list(opciones_sugeridas.keys()), key="select_sugerencia_predictiva")
            
            if st.button(f"➕ Agregar selección actual al carrito", key="btn_add_desde_desplegable", type="primary"):
                cod_elegido = opciones_sugeridas[prod_elegido_label]
                agregar_producto_por_codigo(cod_elegido, cant_venta)
                st.rerun()
        else:
            st.warning(f"⚠️ No se encontraron productos con '{busqueda}'")

    # BOTONES DE ACCIÓN RÁPIDA
    sub_nav = st.columns(4)
    btn_ins = sub_nav[0].button("📌 Artículo Varios", key="btn_ins_v")
    btn_pendiente = sub_nav[1].button("⏳ Dejar en Espera (Pendiente)", key="btn_pend_v")
    btn_recuperar = sub_nav[2].button(f"📂 Recuperar Pendientes ({len(st.session_state.tickets_pendientes)})", key="btn_rec_v")
    btn_del = sub_nav[3].button("🗑️ Borrar Carrito Actual", key="btn_del_v")

    if btn_ins:
        st.session_state.carrito.append({"codigo": "VAR-001", "descripcion": "Artículo General / Varios", "precio": 5.00, "costo": 3.00, "cant": cant_venta, "importe": cant_venta * 5.00, "costo_subtotal": cant_venta * 3.00, "existencia": 999.0})
        st.session_state.mensaje_cobro = None
        st.rerun()

    if btn_pendiente:
        if st.session_state.carrito:
            st.session_state.tickets_pendientes.append(st.session_state.carrito)
            st.session_state.carrito = []
            st.session_state.mensaje_cobro = "⏳ Carrito guardado en espera (pendiente) correctamente."
            st.rerun()
        else:
            st.warning("⚠️ El carrito está vacío para ponerlo en espera.")

    if btn_recuperar:
        if st.session_state.tickets_pendientes:
            st.session_state.carrito = st.session_state.tickets_pendientes.pop(0)
            st.session_state.mensaje_cobro = "📂 Ticket en espera recuperado con éxito."
            st.rerun()
        else:
            st.info("No hay tickets pendientes guardados.")

    if btn_del:
        st.session_state.carrito = []
        st.session_state.mensaje_cobro = None
        st.rerun()

    st.write("---")

    if st.session_state.carrito:
        st.markdown("### 🛒 Productos en el Carrito Actual")
        para_eliminar = None
        for idx, item in enumerate(st.session_state.carrito):
            c1, c2, c3, c4, c5 = st.columns([3, 1.5, 1.5, 1.5, 1])
            c1.write(f"**{item['descripcion']}** (`{item['codigo']}`)")
            c2.write(f"Q{item['precio']:.2f}")
            nueva_cant = c3.number_input("Cantidad", min_value=0.01, value=float(item['cant']), step=0.25, key=f"cant_cart_{idx}_{item['codigo']}")
            if nueva_cant != item['cant']:
                item['cant'] = nueva_cant
                item['importe'] = nueva_cant * item['precio']
                item['costo_subtotal'] = nueva_cant * item.get('costo', 0.0)
                st.session_state.mensaje_cobro = None
                st.rerun()

            c4.write(f"**Q{item['importe']:.2f}**")
            if c5.button("❌", key=f"del_cart_{idx}_{item['codigo']}"):
                para_eliminar = idx

        if para_eliminar is not None:
            st.session_state.carrito.pop(para_eliminar)
            st.session_state.mensaje_cobro = None
            st.rerun()
    else:
        st.info("El carrito está vacío. Indica la cantidad arriba y escribe las primeras letras de un artículo.")

    total_items = sum(item['cant'] for item in st.session_state.carrito)
    total_quetzales = sum(item['importe'] for item in st.session_state.carrito)
    costo_acumulado = sum(item.get('costo_subtotal', 0.0) for item in st.session_state.carrito)

    st.caption(f"**Total acumulado:** {total_items:.2f} unidades/artículos en la venta actual.")

    st.markdown("---")
    col_bot1, col_bot2, col_bot3 = st.columns([2, 2, 2])

    with col_bot1:
        st.markdown(f"**Total a Pagar:** Q{total_quetzales:,.2f}\n**Método:** {metodo_pago}")
        efectivo_recibido = 0.0
        cambio_vuelto = 0.0
        if metodo_pago == "Efectivo" and total_quetzales > 0:
            efectivo_recibido = st.number_input("💵 ¿Con cuánto paga el cliente?", min_value=0.0, value=total_quetzales, step=5.00, key="efectivo_cli")
            cambio_vuelto = efectivo_recibido - total_quetzales
            if cambio_vuelto >= 0:
                st.success(f"🪙 **Cambio / Vuelto:** Q{cambio_vuelto:,.2f}")
            else:
                st.error("⚠️ El efectivo recibido es menor al total.")

    with col_bot2:
        btn_cobrar = st.button("💳 COBRAR Y FINALIZAR", type="primary", use_container_width=True, key="btn_cobrar_final")
        if btn_cobrar:
            if total_quetzales > 0:
                if metodo_pago == "Efectivo" and efectivo_recibido < total_quetzales:
                    st.warning("⚠️ El monto recibido del cliente es insuficiente.")
                else:
                    for item in st.session_state.carrito:
                        cod = item['codigo']
                        if cod in inventario_actual:
                            inventario_actual[cod]['existencia'] -= item['cant']

                    ticket_actual = {
                        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "tienda": st.session_state.tienda_activa,
                        "vendedor": st.session_state.usuario_logueado,
                        "pago": metodo_pago,
                        "items": round(total_items, 2),
                        "total": total_quetzales,
                        "costo_total": costo_acumulado,
                        "detalle": list(st.session_state.carrito)
                    }

                    st.session_state.ventas_realizadas.append(ticket_actual)
                    st.session_state.ultimo_ticket = ticket_actual

                    st.session_state.mensaje_cobro = f"✅ ¡Venta cobrada con éxito! Total: Q{total_quetzales:,.2f} | Cambio: Q{cambio_vuelto:,.2f}"
                    st.session_state.carrito = []
                    st.rerun()
            else:
                st.warning("⚠️ Agrega al menos un producto para cobrar.")

    with col_bot3:
        st.markdown(f'<div class="total-display"><span class="total-amount">Q{total_quetzales:,.2f}</span></div>', unsafe_allow_html=True)

    st.write("---")
    col_caj1, col_caj2 = st.columns(2)
    if col_caj1.button("🖨️ Ver / Reimprimir Último Ticket"):
        if st.session_state.ultimo_ticket:
            st.info(f"🧾 **Último Ticket Cobrado:** {st.session_state.ultimo_ticket['fecha']} | Tienda: {st.session_state.ultimo_ticket['tienda']} | Total: Q{st.session_state.ultimo_ticket['total']:,.2f} ({st.session_state.ultimo_ticket['pago']})")
        else:
            st.info("No hay transacciones registradas en esta sesión reciente.")

# --- VISTA: PROVEEDORES ---
elif st.session_state.vista_actual == "proveedores":
    st.subheader("🚛 Registro y Gestión Completa de Proveedores")
    
    if st.session_state.mensaje_prov_accion:
        st.success(st.session_state.mensaje_prov_accion)
        st.session_state.mensaje_prov_accion = None

    tab_pr1, tab_pr2 = st.tabs(["📋 Directorio de Proveedores", "➕ / ✏️ Gestionar Proveedor"])

    with tab_pr1:
        if st.session_state.proveedores:
            df_prov_show = pd.DataFrame(st.session_state.proveedores)[["empresa", "propietario", "telefono", "direccion", "area", "notas"]]
            df_prov_show.columns = ["Empresa", "Propietario / Contacto", "Teléfono", "Dirección", "Área", "Notas"]
            st.dataframe(df_prov_show, use_container_width=True, hide_index=True)
        else:
            st.info("No hay proveedores registrados por el momento.")

    with tab_pr2:
        if not es_admin:
            st.info("ℹ️ La edición o alta de proveedores está reservada para Administradores.")
        else:
            sub_tab1, sub_tab2 = st.tabs(["➕ Registrar Nuevo Proveedor", "✏️ Editar / Borrar Proveedor"])
            
            with sub_tab1:
                with st.form("form_nuevo_prov", clear_on_submit=True):
                    cp1, cp2 = st.columns(2)
                    p_empresa = cp1.text_input("Empresa / Nombre del Proveedor *")
                    p_propietario = cp2.text_input("Nombre del Propietario / Contacto")
                    
                    cp3, cp4 = st.columns(2)
                    p_tel = cp3.text_input("Número de Contacto / Teléfono")
                    p_area = cp4.selectbox("Área que Surtirá:", AREAS_OFICIALES)
                    
                    p_direccion = st.text_input("Dirección Física (Ubicación)")
                    p_notas = st.text_area("Notas, Días de Visita o Crédito (Opcional):")
                    
                    btn_save_prov = st.form_submit_button("💾 Guardar Proveedor", type="primary")

                    if btn_save_prov and p_empresa.strip():
                        nuevo_id = len(st.session_state.proveedores) + 1
                        st.session_state.proveedores.append({
                            "id": nuevo_id,
                            "empresa": p_empresa.strip(),
                            "propietario": p_propietario.strip(),
                            "telefono": p_tel.strip(),
                            "direccion": p_direccion.strip(),
                            "area": p_area,
                            "notas": p_notas.strip()
                        })
                        st.session_state.mensaje_prov_accion = f"✅ Proveedor '{p_empresa}' registrado correctamente."
                        st.rerun()

            with sub_tab2:
                if st.session_state.proveedores:
                    dict_prov = {f"{p['empresa']} ({p['area']})": idx for idx, p in enumerate(st.session_state.proveedores)}
                    sel_prov_str = st.selectbox("Selecciona Proveedor a Modificar:", list(dict_prov.keys()))
                    idx_p = dict_prov[sel_prov_str]
                    p_data = st.session_state.proveedores[idx_p]

                    e_empresa = st.text_input("Empresa:", value=p_data["empresa"], key="ep_emp")
                    e_propietario = st.text_input("Propietario / Contacto:", value=p_data.get("propietario", ""), key="ep_prop")
                    e_tel = st.text_input("Teléfono:", value=p_data["telefono"], key="ep_tel")
                    e_direccion = st.text_input("Dirección:", value=p_data.get("direccion", ""), key="ep_dir")
                    
                    idx_a = AREAS_OFICIALES.index(p_data["area"]) if p_data["area"] in AREAS_OFICIALES else 0
                    e_area = st.selectbox("Área:", AREAS_OFICIALES, index=idx_a, key="ep_area")
                    e_notas = st.text_area("Notas:", value=p_data["notas"], key="ep_not")

                    btn_pcol1, btn_pcol2 = st.columns(2)
                    if btn_pcol1.button("💾 Actualizar Proveedor", type="primary", key="btn_upd_prov"):
                        p_data["empresa"] = e_empresa
                        p_data["propietario"] = e_propietario
                        p_data["telefono"] = e_tel
                        p_data["direccion"] = e_direccion
                        p_data["area"] = e_area
                        p_data["notas"] = e_notas
                        st.session_state.mensaje_prov_accion = "✅ Datos de proveedor actualizados."
                        st.rerun()

                    if btn_pcol2.button("🗑️ Eliminar Proveedor", key="btn_del_prov"):
                        st.session_state.proveedores.pop(idx_p)
                        st.session_state.mensaje_prov_accion = "🗑️ Proveedor eliminado."
                        st.rerun()

# --- VISTA: PRODUCTOS ---
elif st.session_state.vista_actual == "productos":
    st.subheader(f"🏷️ Gestión de Productos - {st.session_state.tienda_activa}")
    
    if st.session_state.mensaje_prod_accion:
        st.success(st.session_state.mensaje_prod_accion)
        st.session_state.mensaje_prod_accion = None

    if st.session_state.rol_logueado != "Administrador":
        st.error("⛔ Acceso denegado: Únicamente un perfil de Administrador puede crear, editar o eliminar artículos.")
    else:
        tab1, tab2 = st.tabs(["➕ Crear Nuevo Producto", "✏️ Modificar / Eliminar Producto Existente"])
        
        nombres_proveedores = [p["empresa"] for p in st.session_state.proveedores] + ["Sin Proveedor", "➕ Registrar nuevo proveedor..."]

        with tab1:
            st.markdown("### Datos del Producto")
            col_f1, col_f2 = st.columns(2)
            with col_f1:
                nuevo_codigo = st.text_input("Código de Barras / Código Manual *", key="n_code", help="Escanea el código de barras o escribe una clave manual.")
                nuevo_nombre = st.text_input("Nombre / Descripción *", key="n_name")
                nueva_area = st.selectbox("Área / Categoría del Producto *", AREAS_OFICIALES, key="n_area")
                
                prov_sel_prod = st.selectbox("Proveedor del Producto:", nombres_proveedores, key="n_prov_sel")
                
                nuevo_prov_extra = ""
                if prov_sel_prod == "➕ Registrar nuevo proveedor...":
                    nuevo_prov_extra = st.text_input("Escribe el nombre del NUEVO Proveedor:", key="n_prov_nuevo_texto")

            with col_f2:
                precio_venta = st.number_input("Precio de Venta (Q) *", min_value=0.0, step=0.50, format="%.2f", key="n_pventa")
                precio_costo = st.number_input("Precio de Costo (Q)", min_value=0.0, step=0.50, format="%.2f", key="n_pcosto", help="Para cálculo de utilidades y ganancias.")
                existencia_inicial = st.number_input("Cantidad Inicial en Stock", min_value=0.0, step=1.0, format="%.2f", key="n_exis")
                stock_minimo = st.number_input("Stock Mínimo (Alerta)", min_value=0.1, value=5.0, step=1.0, key="n_stmin")

            st.markdown("---")
            st.markdown("📷 **Foto del Producto (Opcional):**")
            tipo_foto_opt = st.radio("¿Cómo deseas agregar la foto?", ["Subir archivo", "Tomar foto con la cámara"], horizontal=True, key="modo_foto_radio")
            
            foto_final_bytes = None
            if tipo_foto_opt == "Subir archivo":
                foto_subida = st.file_uploader("Adjuntar archivo de imagen desde tu dispositivo:", type=["jpg", "png", "jpeg"], key="uploader_prod_nuevo")
                if foto_subida is not None:
                    foto_final_bytes = foto_subida.read()
            else:
                foto_camara = st.camera_input("Toma una foto en vivo del producto:")
                if foto_camara is not None:
                    foto_final_bytes = foto_camara.read()

            st.markdown("---")
            if st.button(f"💾 Guardar Producto en {st.session_state.tienda_activa}", type="primary", key="btn_save_prod"):
                if nuevo_codigo.strip() == "" or nuevo_nombre.strip() == "":
                    st.error("❌ El código y el nombre son obligatorios.")
                else:
                    proveedor_final = prov_sel_prod
                    if prov_sel_prod == "➕ Registrar nuevo proveedor...":
                        if nuevo_prov_extra.strip() != "":
                            proveedor_final = nuevo_prov_extra.strip()
                            st.session_state.proveedores.append({
                                "id": len(st.session_state.proveedores) + 1,
                                "empresa": proveedor_final,
                                "propietario": "",
                                "telefono": "",
                                "direccion": "",
                                "area": nueva_area,
                                "notas": "Creado desde alta de producto"
                            })
                        else:
                            proveedor_final = "Sin Proveedor"

                    inventario_actual[nuevo_codigo.strip()] = {
                        "nombre": nuevo_nombre.strip(),
                        "area": nueva_area,
                        "precio": precio_venta,
                        "costo": precio_costo,
                        "existencia": existencia_inicial,
                        "stock_minimo": stock_minimo,
                        "proveedor": proveedor_final,
                        "foto": foto_final_bytes,
                        "creado_por": st.session_state.usuario_logueado,
                        "fecha_registro": datetime.now().strftime("%Y-%m-%d %H:%M")
                    }
                    st.session_state.mensaje_prod_accion = f"✅ ¡Producto '{nuevo_nombre}' guardado con éxito!"
                    st.rerun()

        with tab2:
            codigos_existentes = list(inventario_actual.keys())
            if codigos_existentes:
                opciones_prod = {f"{v['nombre']} ({k})": k for k, v in inventario_actual.items()}
                prod_sel_label = st.selectbox("Selecciona el producto a modificar o eliminar:", list(opciones_prod.keys()), key="sel_edit_p")
                cod_edit = opciones_prod[prod_sel_label]
                prod_data = inventario_actual[cod_edit]
                
                creador = prod_data.get("creado_por", "Desconocido")
                f_reg = prod_data.get("fecha_registro", "Fecha previa")
                st.caption(f"ℹ️ **Auditoría:** Ingresado por **{creador}** el `{f_reg}`")

                col_e1, col_e2 = st.columns([2, 1])
                with col_e2:
                    st.write("**Foto Registrada:**")
                    if prod_data.get("foto"):
                        st.image(prod_data["foto"], width=150)
                    else:
                        st.info("Sin foto guardada")

                with col_e1:
                    edit_nombre = st.text_input("Nombre / Descripción", value=prod_data["nombre"], key="e_nombre")
                    
                    idx_area_actual = AREAS_OFICIALES.index(prod_data.get("area", AREAS_OFICIALES[0])) if prod_data.get("area") in AREAS_OFICIALES else 0
                    edit_area = st.selectbox("Área / Categoría:", AREAS_OFICIALES, index=idx_area_actual, key="e_area")
                    
                    prov_actual_prod = prod_data.get("proveedor", "Sin Proveedor")
                    lista_provs_edit = [p["empresa"] for p in st.session_state.proveedores] + ["Sin Proveedor"]
                    if prov_actual_prod not in lista_provs_edit:
                        lista_provs_edit.append(prov_actual_prod)
                    
                    idx_prov_edit = lista_provs_edit.index(prov_actual_prod) if prov_actual_prod in lista_provs_edit else 0
                    edit_proveedor = st.selectbox("Proveedor:", lista_provs_edit, index=idx_prov_edit, key="e_prov")

                    edit_precio = st.number_input("Precio Venta (Q)", value=float(prod_data["precio"]), step=0.50, key="e_precio")
                    edit_costo = st.number_input("Precio Costo (Q)", value=float(prod_data.get("costo", 0.0)), step=0.50, key="e_costo")
                    edit_existencia = st.number_input("Existencia actual en stock", value=float(prod_data["existencia"]), step=0.5, key="e_exis")
                    
                    st.markdown("📷 **Actualizar Imagen:**")
                    tipo_foto_edit = st.radio("¿Cómo deseas actualizar la foto?", ["Mantener actual", "Subir archivo", "Tomar foto con la cámara"], horizontal=True, key="modo_foto_edit_radio")
                    
                    nueva_foto_bytes = prod_data.get("foto")
                    if tipo_foto_edit == "Subir archivo":
                        nueva_foto_file = st.file_uploader("Seleccionar nuevo archivo:", type=["jpg", "png", "jpeg"], key="uploader_prod_edit")
                        if nueva_foto_file is not None:
                            nueva_foto_bytes = nueva_foto_file.read()
                    elif tipo_foto_edit == "Tomar foto con la cámara":
                        nueva_foto_cam = st.camera_input("Tomar nueva foto en vivo:")
                        if nueva_foto_cam is not None:
                            nueva_foto_bytes = nueva_foto_cam.read()

                    st.markdown("---")
                    btn_col1, btn_col2 = st.columns(2)
                    
                    with btn_col1:
                        if st.button("💾 Actualizar Cambios", key="btn_update_p", type="primary"):
                            inventario_actual[cod_edit]["nombre"] = edit_nombre
                            inventario_actual[cod_edit]["area"] = edit_area
                            inventario_actual[cod_edit]["proveedor"] = edit_proveedor
                            inventario_actual[cod_edit]["precio"] = edit_precio
                            inventario_actual[cod_edit]["costo"] = edit_costo
                            inventario_actual[cod_edit]["existencia"] = edit_existencia
                            inventario_actual[cod_edit]["foto"] = nueva_foto_bytes
                            inventario_actual[cod_edit]["modificado_por"] = st.session_state.usuario_logueado
                            st.session_state.mensaje_prod_accion = f"✅ Producto '{edit_nombre}' actualizado correctamente."
                            st.rerun()

                    with btn_col2:
                        if st.button("🗑️ ELIMINAR PRODUCTO", key="btn_delete_p"):
                            st.session_state.confirmar_borrado_cod = cod_edit

                    if st.session_state.confirmar_borrado_cod == cod_edit:
                        st.warning(f"⚠️ ¿Eliminar permanentemente **'{prod_data['nombre']}'**?")
                        c_conf1, c_conf2 = st.columns(2)
                        if c_conf1.button("✔️ Sí, Eliminar Definitivamente", key="btn_confirm_del_yes"):
                            del inventario_actual[cod_edit]
                            st.session_state.confirmar_borrado_cod = None
                            st.session_state.mensaje_prod_accion = f"🗑️ Producto eliminado del inventario por {st.session_state.usuario_logueado}."
                            st.rerun()
                        if c_conf2.button("❌ Cancelar", key="btn_confirm_del_no"):
                            st.session_state.confirmar_borrado_cod = None
                            st.rerun()
            else:
                st.info("No hay productos cargados en esta sucursal.")

# --- VISTA: INVENTARIO (Con fotos, entradas/salidas y reporte Excel con logo pequeño y elegante) ---
elif st.session_state.vista_actual == "inventario":
    st.subheader(f"📦 Inventario y Movimientos - {st.session_state.tienda_activa}")
    
    if es_admin:
        tab_inv1, tab_inv2, tab_inv3 = st.tabs(["📋 Existencias Actuales", "📥 Entradas / 📤 Salidas", "📥 Descargar Reporte Excel"])
    else:
        tab_inv1, tab_inv2, tab_inv3 = st.tabs(["📋 Consulta de Existencias", "🔒 Restringido", "📥 Descargar Reporte Excel"])

    with tab_inv1:
        st.markdown("#### 📋 Listado de Existencias y Fotografías")
        if inventario_actual:
            for cod, info in inventario_actual.items():
                es_alerta = "⚠️ STOCK BAJO" if info.get('existencia', 0) <= info.get('stock_minimo', 0) else "✅ OK"
                
                with st.container():
                    col_img, col_info1, col_info2, col_info3 = st.columns([1, 2.5, 1.5, 1.5])
                    
                    with col_img:
                        if info.get("foto"):
                            st.image(info["foto"], width=90)
                        else:
                            st.caption("📷 Sin Foto")
                            
                    with col_info1:
                        st.markdown(f"**{info.get('nombre')}**")
                        st.caption(f"Código: `{cod}` | Área: *{info.get('area')}*")
                        if es_admin:
                            st.caption(f"Proveedor: {info.get('proveedor', 'Sin Proveedor')}")
                            
                    with col_info2:
                        st.markdown(f"**Precio:** Q{info.get('precio', 0.0):,.2f}")
                        if es_admin:
                            st.caption(f"Costo: Q{info.get('costo', 0.0):,.2f}")
                            
                    with col_info3:
                        st.markdown(f"**Stock:** {info.get('existencia', 0.0):,.2f}")
                        st.markdown(f"*{es_alerta}*")
                        
                    st.divider()
        else:
            st.info("No hay productos en inventario.")

    with tab_inv2:
        if es_admin:
            st.markdown("#### Registro de Entradas (Compras/Devoluciones) y Salidas (Mermas/Ajustes)")
            if inventario_actual:
                opciones_inv = {f"{v['nombre']} (Stock actual: {v['existencia']:.1f})": k for k, v in inventario_actual.items()}
                prod_sel_Mov = st.selectbox("Selecciona el Producto:", list(opciones_inv.keys()), key="sel_mov_prod")
                cod_mov = opciones_inv[prod_sel_Mov]

                tipo_mov = st.radio("Tipo de Movimiento:", ["📥 Entrada de Mercadería (Suma al stock)", "📤 Salida / Merma (Resta al stock)"], horizontal=True)
                cant_mov = st.number_input("Cantidad:", min_value=0.1, value=1.0, step=1.0, key="cant_mov_input")
                motivo_mov = st.text_input("Motivo o No. de Factura/Nota:", placeholder="Ej. Compra a proveedor, daño, etc.")

                if st.button("💾 Registrar Movimiento en Inventario", type="primary"):
                    if tipo_mov.startswith("📥"):
                        inventario_actual[cod_mov]["existencia"] += cant_mov
                        accion_txt = "Entrada"
                    else:
                        inventario_actual[cod_mov]["existencia"] -= cant_mov
                        accion_txt = "Salida"

                    st.session_state.movimientos_inventario.append({
                        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "tienda": st.session_state.tienda_activa,
                        "usuario": st.session_state.usuario_logueado,
                        "producto": inventario_actual[cod_mov]["nombre"],
                        "tipo": accion_txt,
                        "cantidad": cant_mov,
                        "motivo": motivo_mov
                    })
                    st.success(f"✅ ¡Movimiento de {accion_txt} registrado con éxito! Nuevo stock: {inventario_actual[cod_mov]['existencia']:.2f}")
                    st.rerun()

                if st.session_state.movimientos_inventario:
                    st.markdown("---")
                    st.markdown("##### Historial de Movimientos Recientes")
                    st.dataframe(pd.DataFrame(st.session_state.movimientos_inventario), use_container_width=True)
            else:
                st.info("No hay productos disponibles para registrar movimientos.")
        else:
            st.warning("⛔ **Acceso Restringido:** Solo los usuarios con rol de Administrador pueden realizar entradas, salidas o ajustes manuales en el inventario.")

    with tab_inv3:
        st.markdown("#### 📥 Generar y Descargar Reporte de Inventario en Excel")
        st.markdown("Filtra los datos por área o proveedor antes de descargar tu reporte profesional.")

        lista_proveedores_disponibles = list(set([info.get("proveedor", "Sin Proveedor") for info in inventario_actual.values()]))
        
        col_f1, col_f2 = st.columns(2)
        filtro_area_excel = col_f1.multiselect("Filtrar por Área(s):", options=AREAS_OFICIALES, default=AREAS_OFICIALES, key="excel_filtro_area")
        filtro_prov_excel = col_f2.multiselect("Filtrar por Proveedor(es):", options=lista_proveedores_disponibles, default=lista_proveedores_disponibles, key="excel_filtro_prov")

        filas_reporte = []
        for cod, info in inventario_actual.items():
            if info.get("area") in filtro_area_excel and info.get("proveedor", "Sin Proveedor") in filtro_prov_excel:
                filas_reporte.append({
                    "Código": cod,
                    "Descripción": info.get("nombre", ""),
                    "Área": info.get("area", ""),
                    "Proveedor": info.get("proveedor", "Sin Proveedor"),
                    "Precio Venta (Q)": float(info.get("precio", 0.0)),
                    "Precio Costo (Q)": float(info.get("costo", 0.0)),
                    "Existencia": float(info.get("existencia", 0.0)),
                    "Stock Mínimo": float(info.get("stock_minimo", 0.0)),
                    "Estado": "⚠️ BAJO" if info.get('existencia', 0) <= info.get('stock_minimo', 0) else "OK"
                })

        if filas_reporte:
            st.markdown(f"📄 **Artículos listos para exportar:** {len(filas_reporte)} producto(s).")

            def generar_excel_inventario_general():
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Inventario General"
                ws.views.sheetView[0].showGridLines = True

                start_row = 6
                
                # Insertar el logo pequeño y elegante en la esquina superior izquierda si existe
                if os.path.exists('logo.png'):
                    try:
                        img_logo = OpenPyxlImage('logo.png')
                        img_logo.width = 110  # Tamaño pequeño y elegante
                        img_logo.height = 40
                        ws.add_image(img_logo, 'A1')
                        ws.row_dimensions[1].height = 35
                    except:
                        pass

                # Nombre de la tienda y subtítulo bien ubicados
                ws["C1"] = f"🏪 {st.session_state.tienda_activa.upper()}"
                ws["C1"].font = Font(name="Calibri", size=13, bold=True, color="1F4E78")
                
                ws["C2"] = "REPORTE GENERAL DE INVENTARIO"
                ws["C2"].font = Font(name="Calibri", size=11, bold=True, color="595959")

                ws["C3"] = f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %I:%M %p')}"
                ws["C3"].font = Font(name="Calibri", size=9, italic=True)

                headers = ["Código", "Descripción", "Área", "Proveedor", "Precio Venta (Q)", "Precio Costo (Q)", "Existencia", "Stock Mínimo", "Estado"]
                for col_idx, h in enumerate(headers, 1):
                    ws.cell(row=start_row, column=col_idx, value=h)

                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 18
                ws.column_dimensions['D'].width = 25
                ws.column_dimensions['E'].width = 18
                ws.column_dimensions['F'].width = 18
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 15
                ws.column_dimensions['I'].width = 15

                for idx, prod in enumerate(filas_reporte, start=start_row + 1):
                    ws.row_dimensions[idx].height = 22
                    ws.cell(row=idx, column=1, value=prod["Código"])
                    ws.cell(row=idx, column=2, value=prod["Descripción"])
                    ws.cell(row=idx, column=3, value=prod["Área"])
                    ws.cell(row=idx, column=4, value=prod["Proveedor"])
                    
                    ce = ws.cell(row=idx, column=5, value=prod["Precio Venta (Q)"])
                    ce.number_format = '"Q "#,##0.00'
                    
                    co = ws.cell(row=idx, column=6, value=prod["Precio Costo (Q)"])
                    co.number_format = '"Q "#,##0.00'
                    
                    ex = ws.cell(row=idx, column=7, value=prod["Existencia"])
                    ex.number_format = '#,##0.00'
                    
                    sm = ws.cell(row=idx, column=8, value=prod["Stock Mínimo"])
                    sm.number_format = '#,##0.00'
                    
                    ws.cell(row=idx, column=9, value=prod["Estado"])

                row_end = start_row + len(filas_reporte)
                tab = Table(displayName="TablaInventarioTotal", ref=f"A{start_row}:I{row_end}")
                tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                ws.add_table(tab)

                output = io.BytesIO()
                wb.save(output)
                return output.getvalue()

            st.download_button(
                label="📥 Descargar Reporte Filtrado en EXCEL (.xlsx)",
                data=generar_excel_inventario_general(),
                file_name=f"Inventario_{st.session_state.tienda_activa.split('(')[0].strip()}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        else:
            st.warning("⚠️ No hay productos que coincidan con los filtros seleccionados.")

# --- VISTA: CLIENTES Y CATÁLOGOS ---
elif st.session_state.vista_actual == "clientes":
    st.subheader("👥 Catálogo Profesional de Productos por Área")
    
    areas_seleccionadas = st.multiselect(
        "Selecciona las Áreas que deseas incluir en el catálogo del cliente:",
        options=AREAS_OFICIALES,
        default=AREAS_OFICIALES
    )

    if areas_seleccionadas:
        prods_filtrados = [
            {
                "Código": cod,
                "Producto": info["nombre"],
                "Área": info["area"],
                "Precio Venta": float(info["precio"]),
                "foto": info.get("foto")
            }
            for cod, info in inventario_actual.items()
            if info["area"] in areas_seleccionadas
        ]

        if prods_filtrados:
            st.markdown(f"#### 📄 Vista previa del catálogo ({len(prods_filtrados)} artículos)")
            
            for p in prods_filtrados:
                cp1, cp2, cp3 = st.columns([1, 3, 1])
                with cp1:
                    if p["foto"]:
                        st.image(p["foto"], width=80)
                    else:
                        st.caption("📷 Sin Foto")
                cp2.write(f"**{p['Producto']}** (`{p['Código']}`)\nÁrea: *{p['Área']}*")
                cp3.write(f"### Q{p['Precio Venta']:.2f}")
                st.divider()

            def generar_excel_con_fotos():
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Catálogo de Productos"
                ws.views.sheetView[0].showGridLines = True

                start_row = 7
                
                if os.path.exists('logo.png'):
                    img_logo = OpenPyxlImage('logo.png')
                    img_logo.width = 160
                    img_logo.height = 60
                    ws.add_image(img_logo, 'A1')
                    ws.row_dimensions[1].height = 50
                    ws["C2"] = f"🏪 {st.session_state.tienda_activa.upper()}"
                    ws["C2"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")

                ws["A5"] = f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %I:%M %p')}"
                ws["A5"].font = Font(name="Calibri", size=10, bold=True)

                headers = ["Foto", "Código", "Producto / Descripción", "Área / Sección", "Precio de Venta (Q)"]
                for col_idx, h in enumerate(headers, 1):
                    ws.cell(row=start_row, column=col_idx, value=h)

                ws.column_dimensions['A'].width = 16
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20

                for idx, prod in enumerate(prods_filtrados, start=start_row + 1):
                    ws.row_dimensions[idx].height = 55
                    
                    if prod["foto"]:
                        try:
                            img_stream = io.BytesIO(prod["foto"])
                            img_p = OpenPyxlImage(img_stream)
                            img_p.width = 60
                            img_p.height = 60
                            ws.add_image(img_p, f"A{idx}")
                        except:
                            ws.cell(row=idx, column=1, value="[Foto]")
                    else:
                        ws.cell(row=idx, column=1, value="Sin Foto")

                    ws.cell(row=idx, column=2, value=prod["Código"])
                    ws.cell(row=idx, column=3, value=prod["Producto"])
                    ws.cell(row=idx, column=4, value=prod["Área"])
                    cp = ws.cell(row=idx, column=5, value=prod["Precio Venta"])
                    cp.number_format = '"Q "#,##0.00'

                row_end = start_row + len(prods_filtrados)
                tab = Table(displayName="TablaCatOficial", ref=f"A{start_row}:E{row_end}")
                tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                ws.add_table(tab)

                output = io.BytesIO()
                wb.save(output)
                return output.getvalue()

            st.download_button(
                label="📥 Descargar Catálogo Profesional por Áreas en EXCEL (.xlsx)",
                data=generar_excel_con_fotos(),
                file_name=f"Catalogo_{st.session_state.tienda_activa.split('(')[0].strip()}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        else:
            st.info("No hay productos registrados en las áreas seleccionadas.")

# --- VISTA: REPORTES DE GANANCIAS Y GRÁFICAS ESTADÍSTICAS ---
elif st.session_state.vista_actual == "ganancias":
    st.subheader(f"📈 Reporte de Ganancias, Utilidades y Gráficas - {st.session_state.tienda_activa}")
    
    if not es_admin:
        st.error("⛔ Acceso Denegado: Módulo exclusivo para Administradores.")
    else:
        tab_g1, tab_g2, tab_g3, tab_g4 = st.tabs(["📅 Ganancias Por Día", "📆 Ganancias Por Quincena", "🗓️ Ganancias Por Mes", "📊 Gráficas y Estadísticas"])

        ventas_t = [v for v in st.session_state.ventas_realizadas if v.get("tienda") == st.session_state.tienda_activa]

        with tab_g1:
            fecha_sel = st.date_input("Selecciona el Día a consultar:", value=datetime.now().date(), key="sel_dia_gan")
            fecha_str = fecha_sel.strftime("%Y-%m-%d")
            
            v_dia = [v for v in ventas_t if v["fecha"].startswith(fecha_str)]
            
            tot_venta_dia = sum(v["total"] for v in v_dia)
            tot_costo_dia = sum(v.get("costo_total", 0.0) for v in v_dia)
            ganancia_dia = tot_venta_dia - tot_costo_dia
            margen_dia = (ganancia_dia / tot_venta_dia * 100) if tot_venta_dia > 0 else 0.0

            col_m1, col_m2, col_m3, col_m4 = st.columns(4)
            col_m1.metric("💵 Ventas del Día", f"Q{tot_venta_dia:,.2f}")
            col_m2.metric("📦 Costo de Mercadería", f"Q{tot_costo_dia:,.2f}")
            col_m3.metric("💰 Ganancia Neta", f"Q{ganancia_dia:,.2f}")
            col_m4.metric("📊 Margen Utilidad", f"{margen_dia:.1f}%")

            if v_dia:
                st.markdown("#### Detalle de Transacciones del Día")
                st.dataframe(pd.DataFrame(v_dia)[["fecha", "vendedor", "pago", "total", "costo_total"]], use_container_width=True)

        with tab_g2:
            col_q1, col_q2 = st.columns(2)
            anio_q = col_q1.number_input("Año:", min_value=2024, max_value=2030, value=datetime.now().year)
            mes_q = col_q2.selectbox("Mes:", ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"], index=datetime.now().month-1)
            num_mes = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"].index(mes_q) + 1

            quincena_sel = st.radio("Selecciona la Quincena:", ["1ª Quincena (Días 1 al 15)", "2ª Quincena (Día 16 al fin de mes)"], horizontal=True)

            v_quin = []
            for v in ventas_t:
                try:
                    f_dt = datetime.strptime(v["fecha"], "%Y-%m-%d %H:%M:%S")
                    if f_dt.year == anio_q and f_dt.month == num_mes:
                        if quincena_sel.startswith("1ª") and 1 <= f_dt.day <= 15:
                            v_quin.append(v)
                        elif quincena_sel.startswith("2ª") and f_dt.day >= 16:
                            v_quin.append(v)
                except:
                    pass

            tot_venta_q = sum(v["total"] for v in v_quin)
            tot_costo_q = sum(v.get("costo_total", 0.0) for v in v_quin)
            ganancia_q = tot_venta_q - tot_costo_q
            margen_q = (ganancia_q / tot_venta_q * 100) if tot_venta_q > 0 else 0.0

            st.markdown("---")
            col_qm1, col_qm2, col_qm3, col_qm4 = st.columns(4)
            col_qm1.metric("💵 Ventas Quincena", f"Q{tot_venta_q:,.2f}")
            col_qm2.metric("📦 Costo Mercadería", f"Q{tot_costo_q:,.2f}")
            col_qm3.metric("💰 Ganancia Neta Quincena", f"Q{ganancia_q:,.2f}")
            col_qm4.metric("📊 Margen Utilidad", f"{margen_q:.1f}%")

            if v_quin:
                st.dataframe(pd.DataFrame(v_quin)[["fecha", "vendedor", "pago", "total", "costo_total"]], use_container_width=True)

        with tab_g3:
            col_m_1, col_m_2 = st.columns(2)
            anio_m = col_m_1.number_input("Año Consulta:", min_value=2024, max_value=2030, value=datetime.now().year, key="anio_m_gan")
            mes_m = col_m_2.selectbox("Mes Consulta:", ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"], index=datetime.now().month-1, key="mes_m_gan")
            num_mes_m = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"].index(mes_m) + 1

            v_mes = []
            for v in ventas_t:
                try:
                    f_dt = datetime.strptime(v["fecha"], "%Y-%m-%d %H:%M:%S")
                    if f_dt.year == anio_m and f_dt.month == num_mes_m:
                        v_mes.append(v)
                except:
                    pass

            tot_venta_m = sum(v["total"] for v in v_mes)
            tot_costo_m = sum(v.get("costo_total", 0.0) for v in v_mes)
            ganancia_m = tot_venta_m - tot_costo_m
            margen_m = (ganancia_m / tot_venta_m * 100) if tot_venta_m > 0 else 0.0

            st.markdown("---")
            col_mm1, col_mm2, col_mm3, col_mm4 = st.columns(4)
            col_mm1.metric("💵 Total Ventas Mes", f"Q{tot_venta_m:,.2f}")
            col_mm2.metric("📦 Costo Total Mes", f"Q{tot_costo_m:,.2f}")
            col_mm3.metric("💰 Ganancia Neta Mensual", f"Q{ganancia_m:,.2f}")
            col_mm4.metric("📊 Margen Utilidad Mensual", f"{margen_m:.1f}%")

        with tab_g4:
            st.markdown("#### 📊 Análisis Gráfico del Comportamiento de Ventas")
            
            if ventas_t:
                df_graf = pd.DataFrame(ventas_t)
                df_graf['fecha_dt'] = pd.to_datetime(df_graf['fecha'])
                df_graf['solo_fecha'] = df_graf['fecha_dt'].dt.strftime('%Y-%m-%d')
                df_graf['mes_str'] = df_graf['fecha_dt'].dt.strftime('%Y-%m')
                df_graf['anio_str'] = df_graf['fecha_dt'].dt.strftime('%Y')

                # Gráfica 1: Venta Diaria
                st.markdown("##### 📅 Comportamiento de Ventas Diarias")
                df_diario = df_graf.groupby('solo_fecha')['total'].sum().reset_index()
                df_diario.columns = ['Fecha', 'Total Vendido (Q)']
                df_diario = df_diario.set_index('Fecha')
                st.bar_chart(df_diario)

                st.markdown("---")

                # Gráfica 2: Venta Mensual
                st.markdown("##### 🗓️ Tendencia de Ventas Mensuales")
                df_mensual = df_graf.groupby('mes_str')['total'].sum().reset_index()
                df_mensual.columns = ['Mes', 'Total Vendido (Q)']
                df_mensual = df_mensual.set_index('Mes')
                st.line_chart(df_mensual)

                st.markdown("---")

                # Gráfica 3: Venta Anual
                st.markdown("##### 📈 Acumulado Anual de Ventas")
                df_anual = df_graf.groupby('anio_str')['total'].sum().reset_index()
                df_anual.columns = ['Año', 'Total Vendido (Q)']
                df_anual = df_anual.set_index('Año')
                st.bar_chart(df_anual)
            else:
                st.info("Aún no hay suficientes registros de ventas en esta tienda para generar las gráficas.")

# --- VISTA: CONFIGURACIÓN DE USUARIOS ---
elif st.session_state.vista_actual == "config":
    st.subheader("⚙️ Configuración y Gestión de Usuarios por Tienda")
    
    if st.session_state.mensaje_user_accion:
        st.success(st.session_state.mensaje_user_accion)
        st.session_state.mensaje_user_accion = None

    if st.session_state.rol_logueado != "Administrador":
        st.error("⛔ Acceso denegado: Únicamente un perfil de Administrador puede gestionar usuarios.")
    else:
        tab_u1, tab_u2, tab_u3 = st.tabs(["➕ Crear Nuevo Usuario", "✏️ Modificar / Ver Claves", "🗑️ Eliminar Usuario"])
        
        with tab_u1:
            with st.form("form_nuevo_usuario", clear_on_submit=True):
                col_u1, col_u2, col_u3, col_u4 = st.columns(4)
                nuevo_user = col_u1.text_input("Nombre de Usuario *")
                nueva_pass = col_u2.text_input("Contraseña *", type="password")
                nuevo_rol = col_u3.selectbox("Rol del Usuario", ["Vendedor de tienda", "Administrador"])
                tienda_asignada = col_u4.selectbox("Tienda Asignada *", TIENDAS + ["Ambas Tiendas"])
                
                btn_crear_user = st.form_submit_button("🔑 Crear Usuario", type="primary")

                if btn_crear_user:
                    if nuevo_user.strip() != "" and nueva_pass.strip() != "":
                        if nuevo_user.strip() in st.session_state.usuarios:
                            st.error("❌ El nombre de usuario ya existe en el sistema.")
                        else:
                            st.session_state.usuarios[nuevo_user.strip()] = {
                                "clave": nueva_pass.strip(),
                                "rol": nuevo_rol,
                                "tienda": "Todas" if nuevo_rol == "Administrador" else tienda_asignada,
                                "creado_por": st.session_state.usuario_logueado
                            }
                            st.session_state.mensaje_user_accion = f"✅ Usuario '{nuevo_user.strip()}' registrado exitosamente."
                            st.rerun()
                    else:
                        st.error("❌ El usuario y la contraseña no pueden estar vacíos.")

        with tab_u2:
            usuarios_lista = list(st.session_state.usuarios.keys())
            user_sel = st.selectbox("Selecciona el usuario a consultar o modificar:", usuarios_lista, key="sel_user_edit")
            
            if user_sel:
                u_info = st.session_state.usuarios[user_sel]
                creador_u = u_info.get("creado_por", "Sistema")
                
                st.info(f"🔑 **Contraseña actual de '{user_sel}':** `{u_info['clave']}`")
                st.caption(f"👤 **Rol:** {u_info['rol']} | Registrado por: **{creador_u}**")

                col_eu1, col_eu2, col_eu3 = st.columns(3)
                edit_clave = col_eu1.text_input("Modificar Contraseña:", value=u_info["clave"], key=f"pass_{user_sel}")
                
                idx_rol = 0 if u_info["rol"] == "Vendedor de tienda" else 1
                edit_rol = col_eu2.selectbox("Rol:", ["Vendedor de tienda", "Administrador"], index=idx_rol, key=f"rol_{user_sel}")
                
                opciones_t = TIENDAS + ["Ambas Tiendas"]
                t_val = u_info.get("tienda", "Todas")
                idx_t = opciones_t.index(t_val) if t_val in opciones_t else 0
                edit_tienda = col_eu3.selectbox("Tienda Asignada:", opciones_t, index=idx_t, key=f"tienda_{user_sel}")

                st.markdown("---")
                if st.button("💾 Actualizar Datos del Usuario", type="primary", key="btn_save_u_changes"):
                    st.session_state.usuarios[user_sel]["clave"] = edit_clave.strip()
                    st.session_state.usuarios[user_sel]["rol"] = edit_rol
                    st.session_state.usuarios[user_sel]["tienda"] = "Todas" if edit_rol == "Administrador" else edit_tienda
                    st.session_state.mensaje_user_accion = f"✅ Usuario '{user_sel}' actualizado correctamente."
                    st.rerun()

        with tab_u3:
            st.markdown("#### 🗑️ Sección de Eliminación de Usuarios")
            usuarios_borrar = [u for u in st.session_state.usuarios.keys()]
            user_a_borrar = st.selectbox("Selecciona el usuario que deseas eliminar permanentemente:", usuarios_borrar, key="sel_user_del")

            if st.button("🗑️ Eliminar Usuario Seleccionado", type="primary", key="btn_ejecutar_borrado_user"):
                if user_a_borrar == st.session_state.usuario_logueado:
                    st.error("❌ No puedes eliminar tu propio usuario mientras estás en sesión activa.")
                else:
                    del st.session_state.usuarios[user_a_borrar]
                    st.session_state.mensaje_user_accion = f"🗑️ Usuario '{user_a_borrar}' eliminado del sistema correctamente."
                    st.rerun()

        st.markdown("---")
        st.markdown("### 📋 Lista Completa de Usuarios Registrados y Claves")
        listado_u = []
        for u, d in st.session_state.usuarios.items():
            listado_u.append({
                "Usuario": u,
                "Contraseña": d["clave"],
                "Rol": d["rol"],
                "Tienda Asignada": d.get("tienda", "Todas"),
                "Creado Por": d.get("creado_por", "Sistema")
            })
        st.dataframe(pd.DataFrame(listado_u), use_container_width=True)

# --- CORTE DE CAJA ---
elif st.session_state.vista_actual == "corte":
    st.subheader(f"📊 Corte de Caja - {st.session_state.tienda_activa}")
    ventas_tienda = [v for v in st.session_state.ventas_realizadas if v.get("tienda") == st.session_state.tienda_activa]
    if ventas_tienda:
        df_corte = pd.DataFrame(ventas_tienda)
        col_m1, col_m2, col_m3 = st.columns(3)
        col_m1.metric("💵 Total Efectivo", f"Q{sum(v['total'] for v in ventas_tienda if v.get('pago') == 'Efectivo'):,.2f}")
        col_m2.metric("💳 Total Transferencia / Tarjeta", f"Q{sum(v['total'] for v in ventas_tienda if v.get('pago') in ['Transferencia', 'Tarjeta']):,.2f}")
        col_m3.metric("💰 Total Recaudado del Día", f"Q{sum(v['total'] for v in ventas_tienda):,.2f}")
        st.dataframe(df_corte, use_container_width=True)
    else:
        st.info("No hay ventas registradas hoy en esta sucursal.")

# Pie de página
st.markdown("---")
col_foot1, col_foot2 = st.columns([3, 1])
with col_foot1: st.caption(f"ℹ️ Punto de Venta operando en: **{st.session_state.tienda_activa}**")
with col_foot2: st.caption(f"🕒 {datetime.now().strftime('%d-%b. %I:%M %p')}")